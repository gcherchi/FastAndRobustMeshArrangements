# This script checks the validity of the surface mapping computed with our surface mapping check.
# Our run_surface_mapping.py script creates the folders "vol_in" and "srf_maps". 
# Put this script inside the "vol_in" folder.
# For each "<filename>.mesh" in the "vol_in" folder, the script checks the mapping in the "srf_maps" folder  
# (with name <filename>_smap.txt, as produced from our run_surface_mapping.py script).

import os
import subprocess
import time
import openpyxl
import re
from openpyxl import load_workbook


def sorted_alphanumeric(data):
    convert = lambda text: int(text) if text.isdigit() else text.lower()
    alphanum_key = lambda key: [convert(c) for c in re.split('([0-9]+)', key) ]
    return sorted(data, key=alphanum_key)



def completion_bar(iterations, total):
    # Calculate the completion ratio
    ratio = iterations / total

    # Determine the length of the progress bar
    bar_length = 50
    progress = int(bar_length * ratio)

    # Create the completion bar
    bar = "[" + "#" * progress + "-" * (bar_length - progress) + "]"

    return bar

def clean_folder(folder_path):
    # Iterate over the contents of the folder
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        # Check if the path is a file or directory
        if os.path.isfile(file_path):
            # If it's a file, remove it
            os.remove(file_path)
        elif os.path.isdir(file_path):
            # If it's a directory, recursively clean it
            clean_folder(file_path)
            # Remove the empty directory
            os.rmdir(file_path)


def compile_and_run_cpp(mesh_file):
    executable = os.path.join('cmake-build-debug', 'mesh_arrangement')
    #measure the time it takes to run the process
    start_time_float = time.time()
    
    run_process = subprocess.run([executable, mesh_file], capture_output=True, text=True)
    
    end_time = time.time() - start_time_float

    # Convert elapsed time to hours, minutes, and seconds
    hours = int(end_time // 3600)
    minutes = int((end_time % 3600) // 60)
    seconds = int(end_time % 60)
    milliseconds = int((end_time - int(end_time)) * 1000)

    # Format the elapsed time
    end_time_formatted = "{:02d}:{:02d}:{:02d}.{:03d}".format(hours, minutes, seconds, milliseconds)

    #print(run_process.stdout, "Time: ", end_time_formatted)
    
    #if run_process.stderr:
        #print("Error:", run_process.stderr)

    #remove from the mesh_file the path and the off extension
    mesh_file = mesh_file.replace("./data/test/", "")
    if mesh_file.endswith(".off"):
        mesh_file = mesh_file.replace(".off", "")
    elif mesh_file.endswith(".stl"):
        mesh_file = mesh_file.replace(".stl", "")

    #if the process failed, return the error message
    if run_process.stderr:
        #print("Failed on file:", mesh_file, run_process.stderr)
        result = [mesh_file, "Failed", end_time_formatted, run_process.stderr]
    else:
        #print("Success on file:", mesh_file)
        result = [mesh_file, "Passed", end_time_formatted, ""]
    

    write_to_excel(result, "test_results.xlsx")


def create_excel_file(excel_file):

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Test Results"
    column_headers = ["File Tested", "Run Status", "Time Completation (hh:mm:ss:ms)","Error Message"]

    # Writing headers
    for col_index, value in enumerate(column_headers, start=2):
        sheet.cell(row=2, column=col_index).value = value
        sheet.cell(row=2, column=col_index).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
        sheet.cell(row=1, column=col_index).font = openpyxl.styles.Font(bold=True)

    #center the text in the cells 3 and 4
    for i in range(3, len(column_headers)+2):
        sheet.cell(row=2, column=i).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)


    wb.save(excel_file)


def write_to_excel(output_data, excel_file):
    wb_add = load_workbook(excel_file)

    
    sheet = wb_add["Test Results"]
    ins_row = str(len(sheet['B']) + 1 )
    ins_row = int(ins_row)
    row_index = ins_row
   
    for col_index, value in enumerate(output_data, start=2):
        sheet.cell(row=row_index, column=col_index).value = value
        sheet.column_dimensions[sheet.cell(row=row_index, column=col_index).column_letter].auto_size = True
        sheet.cell(row=row_index, column=col_index).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
        if row_index != 2 and col_index == 3 : #if the value is "Failed" make the row red from the second column to the last
            if value == "Failed":
                for i in range(2, len(output_data)+2):
                    sheet.cell(row=row_index, column=i).fill = openpyxl.styles.PatternFill(start_color="FF4F47", end_color="FF4F47", fill_type = "solid")
                    sheet.cell(row=row_index, column=i).font = openpyxl.styles.Font(color="FFFFFF")
            else:
                for i in range(2, len(output_data)+2):
                    sheet.cell(row=row_index, column=i).fill = openpyxl.styles.PatternFill(start_color="D1FFBD", end_color="D1FFBD", fill_type = "solid")


    # Resize each column to fit the content
    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width


    #center the text in the cells 3 and 4
    for i in range(3, len(output_data)+2):
        sheet.cell(row=row_index, column=i).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Save the Excel file
    wb_add.save(excel_file)

if __name__ == "__main__":

    #clean the terminal screen
    os.system('cls' if os.name == 'nt' else 'clear')

    folder_to_clean = "./results"
    clean_folder(folder_to_clean)

    #if the file exists, delete it
    if os.path.exists("test_results.xlsx"):
        os.remove("test_results.xlsx")

    #if the file does not exist, create it
    if not os.path.exists("test_results.xlsx"):
        create_excel_file("test_results.xlsx")

    #order the files in the folder by name and iterate over them
    results = []
    files = sorted_alphanumeric(os.listdir("./data/test"))

    
    total_files = len(files)-1
    for i, filename in enumerate(files):
        if filename.endswith(".off") or filename.endswith(".stl"):
            progress_bar = completion_bar(i, total_files)
            print("\rProgress: {}% {}".format(int((i / total_files) * 100), progress_bar), end="")
        
            mesh_file = os.path.join("./data/test", filename)
            result = compile_and_run_cpp(mesh_file)
            results.append(result)
       


    

    
