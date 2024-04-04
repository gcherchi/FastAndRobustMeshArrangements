from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Sample data
data = [
    ["Name", "Age", "Gender"],
    ["John", 30, "Male"],
    ["Alice", 25, "Female"],
    ["Bob", 35, "Male"]
]

# Create a new workbook
wb = Workbook()

# Select active worksheet
ws = wb.active

# Add data to worksheet
for row in data:
    ws.append(row)

# Create Excel Table
table = Table(displayName="Table1", ref=f"A1:{get_column_letter(len(data[0]))}{len(data)}")

# Add a table style
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                      showLastColumn=False, showRowStripes=True, showColumnStripes=False)

# Add the table to the worksheet
ws.add_table(table)

# Save the workbook
wb.save("formatted_table.xlsx")
