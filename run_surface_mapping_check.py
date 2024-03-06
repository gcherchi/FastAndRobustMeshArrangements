# This script checks the validity of the surface mapping computed with our surface mapping check.
# Our run_surface_mapping.py script creates the folders "vol_in" and "srf_maps". 
# Put this script inside the "vol_in" folder.
# For each "<filename>.mesh" in the "vol_in" folder, the script checks the mapping in the "srf_maps" folder  
# (with name <filename>_smap.txt, as produced from our run_surface_mapping.py script).

import os
import subprocess
import openpyxl
import re


def sorted_alphanumeric(data):
    convert = lambda text: int(text) if text.isdigit() else text.lower()
    alphanum_key = lambda key: [convert(c) for c in re.split('([0-9]+)', key) ]
    return sorted(data, key=alphanum_key)



def clean_folder(folder_path):
    # Iterate over the contents of the folder
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        # Check if the path is a file or directory
        if os.path.isfile(file_path):
            # If it's a file, remove it
            os.remove(file_path)
        elif os.path.isdir(file_path):
            # If it's a directory, recursively clean it
            clean_folder(file_path)
            # Remove the empty directory
            os.rmdir(file_path)




def compile_and_run_cpp(mesh_file):

    executable = "/Users/michele/Documents/GitHub/FastAndRobustMeshArrangements/cmake-build-debug/mesh_arrangement"
    #executable = os.path.join('cmake-build-debug', 'mesh_arrangement')
    run_process = subprocess.run([executable, mesh_file], capture_output=True, text=True)
    print(run_process.stdout)
    if run_process.stderr:
        print("Error:", run_process.stderr)

    #remove from the mesh_file the path and the off extension
    mesh_file = mesh_file.replace("./data/test/", "")
    if mesh_file.endswith(".off"):
        mesh_file = mesh_file.replace(".off", "")
    elif mesh_file.endswith(".stl"):
        mesh_file = mesh_file.replace(".stl", "")

    #if the process failed, return the error message
    if run_process.stderr:
        print("Failed on file:", mesh_file, run_process.stderr)
        return [mesh_file, "Failed", run_process.stderr]
    else:
        print("Success on file:", mesh_file)
        return [mesh_file, "Passed", ""]


def create_excel_file():

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Test Results"
    row = [2]
    column_headers = ["File Tested", "Run Status", "Error Message"]
    # appy the style to the headers
    for row_index, row_data in enumerate(row, start=2):
        for col_index, value in enumerate(column_headers, start=2):
            sheet.cell(row=row_index, column=col_index).value = value
            sheet.cell(row=row_index, column=col_index).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            sheet.column_dimensions[sheet.cell(row=row_index, column=col_index).column_letter].auto_size = True
            sheet.cell(row=row_index, column=col_index).font = openpyxl.styles.Font(bold=True)

    # Resize each column to fit the content
    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    # Save the Excel file
    wb.save("test_results.xlsx")
    if os.path.exists("test_results.xlsx"):
        print("Excel file 'test_results.xlsx' created successfully.")


def write_to_excel(output_data, excel_file):

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Test Results"

    # Writing data
    for row_index, row_data in enumerate(output_data, start=2):
        for col_index, value in enumerate(row_data, start=2):
            sheet.cell(row=row_index, column=col_index).value = value
            sheet.column_dimensions[sheet.cell(row=row_index, column=col_index).column_letter].auto_size = True
            sheet.cell(row=row_index, column=col_index).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
            if row_index != 2 and col_index == 3 : #if the value is "Failed" make the row red from the second column to the last
                if value == "Failed":
                    for i in range(2, len(row_data)+2):
                        sheet.cell(row=row_index, column=i).fill = openpyxl.styles.PatternFill(start_color="FF4F47", end_color="FF4F47", fill_type = "solid")
                        sheet.cell(row=row_index, column=i).font = openpyxl.styles.Font(color="FFFFFF")
                else:
                    for i in range(2, len(row_data)+2):
                        sheet.cell(row=row_index, column=i).fill = openpyxl.styles.PatternFill(start_color="D1FFBD", end_color="D1FFBD", fill_type = "solid")

    # Resize each column to fit the content
    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    # Save the Excel file
    wb.save(excel_file)
    print(f"Excel file '{excel_file}' updated successfully.")

if __name__ == "__main__":

    folder_to_clean = "./results"
    clean_folder(folder_to_clean)

    #order the files in the folder by name and iterate over them
    results = []
    files = sorted_alphanumeric(os.listdir("./data/test"))
    print(files)

    for filename in files:
       if filename.endswith(".off") or filename.endswith(".stl"):
           mesh_file = os.path.join("./data/test", filename)
           result = compile_and_run_cpp(mesh_file)
           results.append(result)

    results.insert(0, ["File Tested", "Run Status", "Error Message"]);

    #if the file does not exist, create it
    if not os.path.exists("test_results.xlsx"):
        create_excel_file()

    write_to_excel(results, "test_results.xlsx")
